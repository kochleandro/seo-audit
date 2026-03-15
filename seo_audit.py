import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import pandas as pd
import requests
from bs4 import BeautifulSoup
import os
import time
from urllib.parse import urljoin, urlparse
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment

session = requests.Session()

HEADERS = {"User-Agent": "Mozilla/5.0"}

urls_rastreadas = []

# =========================
# LEER URLS DESDE TXT
# =========================

def leer_urls_txt(ruta):

    urls = []

    with open(ruta,"r",encoding="utf-8") as f:

        for linea in f:
            url = linea.strip()

            if url:
                urls.append(url)

    return urls


# =========================
# DETECTAR SITEMAP
# =========================

def detectar_sitemap(dominio):

    posibles = [
        "/sitemap_index.xml",
        "/wp-sitemap.xml",
        "/sitemap.xml"
    ]

    if not dominio.startswith("http"):
        dominio = "https://" + dominio

    dominio = dominio.rstrip("/")

    for ruta in posibles:

        url = dominio + ruta

        try:

            r = session.get(url,headers=HEADERS,timeout=10)

            if r.status_code == 200 and "<loc>" in r.text:
                return url

        except:
            pass

    return None


# =========================
# LEER SITEMAP
# =========================

def obtener_urls_desde_sitemap(sitemap):

    urls = []

    r = session.get(sitemap,headers=HEADERS)

    soup = BeautifulSoup(r.content,"xml")

    sitemaps = soup.find_all("sitemap")

    if sitemaps:

        for sm in sitemaps:

            loc = sm.find("loc")

            if loc:

                sub = loc.text.strip()

                r2 = session.get(sub,headers=HEADERS)

                soup2 = BeautifulSoup(r2.content,"xml")

                for url in soup2.find_all("loc"):
                    urls.append(url.text.strip())

    else:

        for url in soup.find_all("loc"):
            urls.append(url.text.strip())

    return urls


# =========================
# CRAWLER COMPLETO
# =========================

def rastreo_completo():

    global urls_rastreadas

    dominio = entry_archivo.get().strip()

    if not dominio:
        messagebox.showerror("Error","Ingresa dominio")
        return

    if not dominio.startswith("http"):
        dominio = "https://" + dominio

    dominio_base = urlparse(dominio).netloc

    visitadas = set()
    pendientes = [dominio]

    urls_rastreadas = []

    log_text.delete(1.0,tk.END)
    log_text.insert(tk.END,"Iniciando rastreo completo...\n")

    while pendientes and len(urls_rastreadas) < 500:

        url = pendientes.pop(0)

        if url in visitadas:
            continue

        visitadas.add(url)

        try:

            r = session.get(url,headers=HEADERS,timeout=10)

            if r.status_code != 200:
                continue

            soup = BeautifulSoup(r.text,"html.parser")

            urls_rastreadas.append(url)

            log_text.insert(tk.END,f"Rastreando: {url}\n")
            log_text.see(tk.END)

            for link in soup.find_all("a",href=True):

                href = link["href"]

                full = urljoin(url,href)

                parsed = urlparse(full)

                if parsed.netloc == dominio_base:

                    clean = parsed.scheme + "://" + parsed.netloc + parsed.path

                    if clean not in visitadas:
                        pendientes.append(clean)

        except:
            pass

    log_text.insert(tk.END,f"\nTotal URLs encontradas: {len(urls_rastreadas)}\n")


# =========================
# RASTREAR URLS
# =========================

def rastrear_urls():

    global urls_rastreadas

    origen = entry_archivo.get().strip()

    if not origen:
        messagebox.showerror("Error","Ingresa dominio, sitemap o TXT")
        return

    log_text.delete(1.0,tk.END)

    # Leer TXT
    if origen.endswith(".txt"):

        urls_rastreadas = leer_urls_txt(origen)

        log_text.insert(tk.END,f"URLs cargadas desde TXT: {len(urls_rastreadas)}\n")

        return

    if not origen.startswith("http"):
        origen="https://"+origen

    sitemap = detectar_sitemap(origen)

    if sitemap:

        log_text.insert(tk.END,f"Sitemap detectado: {sitemap}\n")

        urls_rastreadas = obtener_urls_desde_sitemap(sitemap)

    else:

        urls_rastreadas = obtener_urls_desde_sitemap(origen)

    log_text.insert(tk.END,f"\nURLs encontradas: {len(urls_rastreadas)}\n")


# =========================
# ANALIZAR URL
# =========================

def analizar_url(url):

    resultado = {
        "URL": url,
        "Status": "",
        "Indexable": "",
        "Title": "",
        "Title_length": 0,
        "MetaDescription": "",
        "MetaDescription_length": 0,
        "H1": "",
        "H1_count": 0,
        "Canonical": "",
        "Images": 0,
        "Images_sin_alt": 0,
        "WordCount": 0,
        "PageSizeKB": 0,
        "BrokenLinks": 0
    }

    try:

        r = session.get(url,headers=HEADERS,timeout=15)

        resultado["Status"] = r.status_code
        resultado["PageSizeKB"] = round(len(r.content)/1024,2)

        if r.status_code == 200:

            soup = BeautifulSoup(r.text,"html.parser")

            if soup.title:
                title = soup.title.get_text(strip=True)
                resultado["Title"] = title
                resultado["Title_length"] = len(title)

            desc = soup.find("meta",attrs={"name":"description"})
            if desc and desc.get("content"):
                d = desc["content"]
                resultado["MetaDescription"] = d
                resultado["MetaDescription_length"] = len(d)

            h1 = soup.find_all("h1")
            resultado["H1_count"] = len(h1)

            if h1:
                resultado["H1"] = h1[0].get_text(strip=True)

            canonical = soup.find("link",rel="canonical")
            if canonical and canonical.get("href"):
                resultado["Canonical"] = canonical["href"]

            imgs = soup.find_all("img")
            resultado["Images"] = len(imgs)

            sin_alt = [img for img in imgs if not img.get("alt")]
            resultado["Images_sin_alt"] = len(sin_alt)

            texto = soup.get_text()
            resultado["WordCount"] = len(texto.split())

            robots = soup.find("meta",attrs={"name":"robots"})
            if robots and "noindex" in robots.get("content","").lower():
                resultado["Indexable"] = "NO"
            else:
                resultado["Indexable"] = "SI"

            # Detectar links rotos
            broken = 0

            for link in soup.find_all("a",href=True):

                href = link["href"]

                if href.startswith("http"):

                    try:

                        r2 = session.head(href,timeout=5)

                        if r2.status_code >= 400:
                            broken += 1

                    except:
                        broken += 1

            resultado["BrokenLinks"] = broken

    except:

        resultado["Status"] = "Error"
        resultado["Indexable"] = "NO"

    time.sleep(0.3)

    return resultado


# =========================
# AUDITORIA
# =========================

def ejecutar_auditoria():

    global urls_rastreadas

    if not urls_rastreadas:
        messagebox.showerror("Error","Primero rastrea URLs")
        return

    resultados = []

    for i,url in enumerate(urls_rastreadas,1):

        log_text.insert(tk.END,f"[{i}/{len(urls_rastreadas)}] Analizando {url}\n")
        log_text.see(tk.END)

        root.update()

        res = analizar_url(url)

        resultados.append(res)

    df = pd.DataFrame(resultados)

    df["Title_duplicado"] = df["Title"].duplicated(keep=False)

    def semaforo(row):

        if row["Status"] != 200:
            return "🔴"

        if row["Indexable"] == "NO":
            return "🔴"

        if row["Title_length"] < 30 or row["Title_length"] > 65:
            return "🟡"

        return "🟢"

    df["Semaforo"] = df.apply(semaforo,axis=1)

    # Nombre del archivo automático

    dominio = urlparse(urls_rastreadas[0]).netloc.replace("www.","")

    fecha = datetime.now().strftime("%Y-%m-%d")

    nombre_archivo = f"auditoria_{dominio}_{fecha}.xlsx"

    archivo = os.path.join(os.getcwd(),nombre_archivo)

    with pd.ExcelWriter(archivo,engine="openpyxl") as writer:

        df.to_excel(writer,index=False,sheet_name="Auditoria")

        sheet = writer.sheets["Auditoria"]

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = "A2"

        for col in sheet.columns:

            max_length = 0
            column = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length,len(str(cell.value)))
                except:
                    pass

            sheet.column_dimensions[column].width = max_length + 3

        verde = PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
        amarillo = PatternFill(start_color="FFF2CC",end_color="FFF2CC",fill_type="solid")
        rojo = PatternFill(start_color="F4CCCC",end_color="F4CCCC",fill_type="solid")

        for row in sheet.iter_rows(min_row=2):

            for cell in row:

                if cell.value == "🟢":
                    cell.fill = verde

                elif cell.value == "🟡":
                    cell.fill = amarillo

                elif cell.value == "🔴":
                    cell.fill = rojo

    messagebox.showinfo("Auditoría completa",f"Archivo generado:\n{archivo}")


# =========================
# GUI
# =========================

def seleccionar_archivo():

    file = filedialog.askopenfilename()

    if file:
        entry_archivo.delete(0,tk.END)
        entry_archivo.insert(0,file)


root = tk.Tk()
root.title("SEO Audit Tool")
root.geometry("760x560")

frame = tk.Frame(root)
frame.pack(pady=10)

tk.Label(frame,text="Dominio / Sitemap / TXT").pack(side=tk.LEFT)

entry_archivo = tk.Entry(frame,width=50)
entry_archivo.pack(side=tk.LEFT,padx=5)

tk.Button(frame,text="Seleccionar",command=seleccionar_archivo).pack(side=tk.LEFT)

tk.Button(root,text="Rastrear URLs (Sitemap o TXT)",command=rastrear_urls,bg="blue",fg="white").pack(pady=5)

tk.Button(root,text="Rastreo Completo (Crawler)",command=rastreo_completo,bg="orange",fg="white").pack(pady=5)

tk.Button(root,text="Ejecutar Auditoría",command=ejecutar_auditoria,bg="green",fg="white").pack(pady=10)

log_text = scrolledtext.ScrolledText(root,width=95,height=24)
log_text.pack(padx=10,pady=10)

root.mainloop()